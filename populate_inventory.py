#!/usr/bin/env python
"""
Script to populate the inventory database from the Equipment List.xlsx fixture.

This script reads data from the Excel file and creates records in the correct
dependency order:
1. Locations
2. DetectorModels
3. SensorTypes
4. DetectorModelConfigurations
5. Detectors
6. Maintenance
7. Sensors

Usage:
    python populate_inventory.py
"""

import os
import sys
import django

# Setup Django environment
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
sys.path.append(BASE_DIR)
os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'apis.settings')
django.setup()

from datetime import datetime, date
from decimal import Decimal
import openpyxl

from inventory.models import (
    Location, LocationType,
    DetectorModel, DetectorType, Manufacturer, Supplier,
    SensorType, Sensor, SensorSlot, SensorGas, SensorStatus,
    DetectorModelConfiguration,
    Detector, DetectorStatus,
    Maintenance, MaintenanceType, MaintenanceStatus,
    MaintenanceTask, MaintenanceTaskType,
)

# Mapping dictionaries to track created objects
locations = {}
detector_models = {}
sensor_types = {}
detector_configurations = {}
detectors = {}


def get_location_type_code(desc):
    """Map location type description to code."""
    mapping = {
        'Station': 'ST',
        'Appliance': 'AP',
        'District Office': 'DI',
        'Work Group': 'WG',
        'Warehouse': 'WA',
        'Workshop': 'WK',
        'External': 'EX',
        'Unknown': 'EX',  # Default unknown to External
        'Offsite Maintenance': 'EX',
    }
    return mapping.get(desc, 'ST')


def get_detector_type_code(desc):
    """Map detector type description to code."""
    mapping = {
        'Personal Multigas Detector': 'NM',
        'Pumped Multigas Detector': 'PM',
        'PID': 'PI',
        'Radio Modem': 'RM',
        'Area Monitor': 'AM',
        'Dosimeter': 'DI',
        'Survey Meter': 'SM',
        'Laptop': 'LT',
        'FID': 'FI',
        'Particulate Counter': 'PC',
        'Area Multigas Monitor': 'AM',
        'Calibration Station': 'CS',
        'IMS': 'IM',
        'FPD': 'FP',
        'Mercury Meter': 'MM',
        'GC/MS': 'GC',
    }
    return mapping.get(desc, 'NM')


def get_manufacturer_code(desc):
    """Map manufacturer description to code."""
    if not desc:
        return None
    mapping = {
        'Honeywell': 'HW',
        'Honeywell/Rae': 'HW',
        'MSA': 'MS',
        'Draeger': 'DR',
        'Proengin': 'PR',
        'Thermo Scientific': 'TH',
        'Smiths': 'SM',
        'Inficon': 'IN',
        'Mirion': 'MI',
        'Health Physics': 'HE',
        'Automess': 'AU',
        'TSI': 'TS',
        'Dell': 'DE',
        'Nippon': 'NI',
        'CAC': 'CA',
        'Clear Gas': 'CG',
    }
    return mapping.get(desc, 'HW')


def get_supplier_code(desc):
    """Map supplier description to code."""
    if not desc:
        return None
    mapping = {
        'AES': 'AE',
        'MSA': 'MS',
        'Draeger': 'DR',
        'Warsach': 'WS',
        'CAC': 'CA',
        'Clear Gas': 'CG',
        'Lear Siegler': 'LS',
        'AirMet': 'AM',
        'Radtronics': 'AE',  # Map to AES as fallback
    }
    return mapping.get(desc, 'AE')


def get_maintenance_type_code(desc):
    """Map maintenance type description to code."""
    mapping = {
        'Routine Service': 'SV',
        'Unscheduled Service': 'MS',
        'Battery Replacement': 'BT',
        'Filter Replacement': 'FC',
        'Dessicant Replacement': 'DR',
        'Clean': 'CL',
    }
    return mapping.get(desc, 'SV')


def get_sensor_gas_code(desc):
    """Map sensor gas description to code."""
    if not desc:
        return None
    mapping = {
        'LEL': 'LE',
        'O2': 'O2',
        'CO': 'CO',
        'HCN': 'HC',
        'H2S': 'HS',
        'NH3': 'NH',
        'Cl2': 'CL',
        'NO2': 'NO',
        'PH3': 'PH',
        'SO2': 'SO',
        'PID': 'VO',  # PID detects VOCs
        'CO/H2S': 'CO',  # Dual sensor, use CO as primary
    }
    return mapping.get(desc, 'VO')


def create_locations(wb):
    """Create all Location records."""
    print("Creating Locations...")
    ws = wb['Location']
    
    created = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        label = row[0]
        if not label:
            continue
            
        station = row[1] or 'N/A'
        address = row[2] or ''
        location_type_desc = row[3]
        priority = row[5] if row[5] else 1
        
        # Map location type
        location_type_code = get_location_type_code(location_type_desc)
        
        location, created_flag = Location.objects.get_or_create(
            label=label,
            defaults={
                'station': station,
                'address': address,
                'location_type': location_type_code,
                'priority': priority,
            }
        )
        locations[label] = location
        if created_flag:
            created += 1

    print(f"  Created {created} new locations (total: {len(locations)})")
    return created


def create_detector_models(wb):
    """Create all DetectorModel records."""
    print("Creating DetectorModels...")
    ws = wb['DetectorModels']
    
    created = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        model_name = row[0]
        if not model_name:
            continue
            
        detector_type_desc = row[1]
        manufacturer_desc = row[3]
        supplier_desc = row[5]
        
        # Map to codes
        detector_type_code = get_detector_type_code(detector_type_desc)
        manufacturer_code = get_manufacturer_code(manufacturer_desc)
        supplier_code = get_supplier_code(supplier_desc)
        
        detector_model, created_flag = DetectorModel.objects.get_or_create(
            label=model_name,
            defaults={
                'detector_type': detector_type_code,
                'manufacturer': manufacturer_code,
                'supplier': supplier_code,
            }
        )
        detector_models[model_name] = detector_model
        if created_flag:
            created += 1
    
    print(f"  Created {created} new detector models (total: {len(detector_models)})")
    return created


def create_sensor_types(wb):
    """Create all SensorType records."""
    print("Creating SensorTypes...")
    ws = wb['SensorTypes']
    
    created = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        sensorgas = row[0]
        part_number = row[1]
        if not part_number:
            continue
            
        manufacturer_desc = row[2]
        compatible_models_str = row[3]
        warranty_months = row[4]
        expiry_months = row[5]
        
        # Convert part_number to string if it's a number
        part_number = str(part_number)
        
        # Map manufacturer
        manufacturer_code = get_manufacturer_code(manufacturer_desc)
        
        # Parse compatible detector models
        compatible_models = []
        if compatible_models_str:
            compatible_models = [m.strip() for m in str(compatible_models_str).split(',')]
        
        # Map sensor gas
        sensor_gas_code = get_sensor_gas_code(sensorgas)
        
        sensor_type, created_flag = SensorType.objects.get_or_create(
            part_number=part_number,
            defaults={
                'sensorgas': sensor_gas_code,
                'manufacturer': manufacturer_code,
                'compatible_detectormodels': ','.join(compatible_models) if compatible_models else '',
                'warranty_months': warranty_months,
                'expiry_months': expiry_months,
                'active': True,
            }
        )
        sensor_types[part_number] = sensor_type
        if created_flag:
            created += 1
    
    print(f"  Created {created} new sensor types (total: {len(sensor_types)})")
    return created


def create_detector_configurations(wb):
    """Create all DetectorModelConfiguration records."""
    print("Creating DetectorModelConfigurations...")
    ws = wb['DetectorConfigurations']

    created = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        label = row[0]
        detector_type_label = row[1]
        sensor_gases_str = row[2]

        if not label or not detector_type_label:
            continue

        # Get the detector model
        if detector_type_label not in detector_models:
            print(f"  Warning: Detector model '{detector_type_label}' not found, skipping configuration '{label}'")
            continue

        detector_model = detector_models[detector_type_label]

        # Parse sensor gases (comma-separated string)
        sensor_gases = ''
        if sensor_gases_str:
            sensor_gases = str(sensor_gases_str)

        config, created_flag = DetectorModelConfiguration.objects.get_or_create(
            detector_model=detector_model,
            label=label,
            defaults={
                'sensor_gases': sensor_gases,
            }
        )
        key = f"{detector_type_label}_{label}"
        detector_configurations[key] = config
        if created_flag:
            created += 1

    print(f"  Created {created} new detector configurations (total: {len(detector_configurations)})")
    return created


def create_detectors(wb):
    """Create all Detector records."""
    print("Creating Detectors...")
    ws = wb['Detectors']

    created = 0
    skipped = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        label = row[0]
        if not label:
            continue

        serial = row[1]
        detector_model_label = row[2]
        location_label = row[3]
        configuration_label = row[4]
        status_desc = row[5]

        # Validate required references
        if detector_model_label not in detector_models:
            print(f"  Warning: Detector model '{detector_model_label}' not found for detector '{label}'")
            skipped += 1
            continue

        if location_label not in locations:
            print(f"  Warning: Location '{location_label}' not found for detector '{label}'")
            skipped += 1
            continue

        detector_model = detector_models[detector_model_label]
        location = locations[location_label]

        # Map status
        status_mapping = {
            'Operational': 'OP',
            'In Stock': 'IS',
            'Offline Repair': 'OF',
            'On Order': 'OO',
            'Decommissioned': 'DC',
        }
        status = status_mapping.get(status_desc, 'OP') if status_desc else 'OP'

        # Find configuration if it exists
        configuration = None
        if configuration_label:
            config_key = f"{detector_model_label}_{configuration_label}"
            if config_key in detector_configurations:
                configuration = detector_configurations[config_key]
            else:
                # Try to find any configuration with this label for this model
                try:
                    configuration = DetectorModelConfiguration.objects.get(
                        detector_model=detector_model,
                        label=configuration_label
                    )
                except DetectorModelConfiguration.DoesNotExist:
                    pass

        detector, created_flag = Detector.objects.get_or_create(
            label=label,
            defaults={
                'serial': serial,
                'detector_model': detector_model,
                'location': location,
                'status': status,
                'configuration': configuration,
            }
        )
        detectors[label] = detector
        if created_flag:
            created += 1

    print(f"  Created {created} new detectors (total: {len(detectors)}, skipped: {skipped})")
    return created


def create_maintenance(wb):
    """Create all Maintenance records."""
    print("Creating Maintenance records...")
    ws = wb['Maintenance']
    
    created = 0
    skipped = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        detector_label = row[0]
        maintenance_type_desc = row[1]
        due_date = row[2]
        
        if not detector_label:
            continue
        
        # Find the detector
        if detector_label not in detectors:
            print(f"  Warning: Detector '{detector_label}' not found for maintenance record")
            skipped += 1
            continue
        
        detector = detectors[detector_label]
        
        # Map maintenance type
        maintenance_type = get_maintenance_type_code(maintenance_type_desc)
        
        # Parse due date
        if due_date and isinstance(due_date, datetime):
            due_date = due_date.date()
        
        maintenance, created_flag = Maintenance.objects.get_or_create(
            detector=detector,
            maintenance_type=maintenance_type,
            date_due=due_date,
            defaults={
                'status': 'SC',  # Scheduled
            }
        )
        if created_flag:
            created += 1
    
    print(f"  Created {created} new maintenance records (skipped: {skipped})")
    return created


def create_sensors(wb):
    """Create all Sensor and SensorSlot records."""
    print("Creating Sensors and SensorSlots...")
    ws = wb['Sensors']

    sensors_created = 0
    slots_updated = 0
    skipped = 0

    for row in ws.iter_rows(min_row=2, values_only=True):
        detector_label = row[0]
        serial = row[1]
        part_number = row[2]
        manufacture_date = row[3]
        warranty_date = row[4]
        expiry_date = row[5]

        if not detector_label or not part_number:
            continue

        # Find the detector
        if detector_label not in detectors:
            print(f"  Warning: Detector '{detector_label}' not found for sensor '{serial}'")
            skipped += 1
            continue

        detector = detectors[detector_label]

        # Convert part_number to string
        part_number = str(part_number)

        # Find the sensor type
        if part_number not in sensor_types:
            print(f"  Warning: Sensor type '{part_number}' not found for sensor '{serial}'")
            skipped += 1
            continue

        sensor_type = sensor_types[part_number]

        # Convert dates if needed
        if manufacture_date and isinstance(manufacture_date, datetime):
            manufacture_date = manufacture_date.date()
        if warranty_date and isinstance(warranty_date, datetime):
            warranty_date = warranty_date.date()
        if expiry_date and isinstance(expiry_date, datetime):
            expiry_date = expiry_date.date()

        # Create sensor
        sensor, created_flag = Sensor.objects.get_or_create(
            serial=serial,
            defaults={
                'sensor_type': sensor_type,
                'status': 'OP',  # Operational
                'receive_date': manufacture_date,
                'warranty_date': warranty_date,
                'expiry_date': expiry_date,
            }
        )
        if created_flag:
            sensors_created += 1

        # Find the sensor slot for this detector by matching sensorgas
        # The sensor slot should have been created by the signals when the detector was created
        try:
            sensor_slot = SensorSlot.objects.get(
                detector=detector,
                sensorgas=sensor_type.sensorgas
            )
            # Update the sensor in the slot
            if sensor_slot.sensor != sensor:
                # If there's already a different sensor in this slot, mark it as decommissioned
                if sensor_slot.sensor:
                    old_sensor = sensor_slot.sensor
                    old_sensor.status = 'DC'
                    old_sensor.remove_date = sensor.install_date or date.today()
                    old_sensor.save()
                sensor_slot.sensor = sensor
                sensor_slot.save()
                slots_updated += 1
        except SensorSlot.DoesNotExist:
            # No slot exists for this sensorgas, create one
            sensor_slot = SensorSlot.objects.create(
                detector=detector,
                sensorgas=sensor_type.sensorgas,
                sensor=sensor
            )
            slots_updated += 1

    print(f"  Created {sensors_created} new sensors, updated {slots_updated} sensor slots (skipped: {skipped})")
    return sensors_created, slots_updated


def main():
    """Main function to populate the database."""
    print("=" * 60)
    print("Populating Inventory Database from Equipment List.xlsx")
    print("=" * 60)
    
    # Load the Excel file
    fixture_path = os.path.join(BASE_DIR, 'inventory', 'fixtures', 'Equipment List.xlsx')
    
    if not os.path.exists(fixture_path):
        print(f"ERROR: Fixture file not found: {fixture_path}")
        return
    
    print(f"\nLoading fixture file: {fixture_path}")
    wb = openpyxl.load_workbook(fixture_path)
    
    # Create records in dependency order
    print("\n--- Phase 1: Base Data ---")
    create_locations(wb)
    create_detector_models(wb)
    create_sensor_types(wb)
    
    print("\n--- Phase 2: Configurations ---")
    create_detector_configurations(wb)
    
    print("\n--- Phase 3: Equipment ---")
    create_detectors(wb)
    
    print("\n--- Phase 4: Maintenance & Sensors ---")
    create_maintenance(wb)
    create_sensors(wb)
    
    print("\n" + "=" * 60)
    print("Database population complete!")
    print("=" * 60)
    print(f"\nSummary:")
    print(f"  Locations: {len(locations)}")
    print(f"  Detector Models: {len(detector_models)}")
    print(f"  Sensor Types: {len(sensor_types)}")
    print(f"  Detector Configurations: {len(detector_configurations)}")
    print(f"  Detectors: {len(detectors)}")


if __name__ == '__main__':
    main()
