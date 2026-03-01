#!/usr/bin/env python
"""
Script to export the inventory database to an Excel spreadsheet.

This script reads data from the database and creates an Excel file in the same
format as the Equipment List.xlsx fixture, with the following sheets:
1. Location
2. DetectorModels
3. SensorTypes
4. DetectorConfigurations
5. Detectors
6. Maintenance
7. Sensors
8. Lookups

Usage:
    python export_inventory.py
"""

import os
import sys
import django

# Setup Django environment
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
sys.path.append(BASE_DIR)
os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'apis.settings')
django.setup()

from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from inventory.models import (
    Location, LocationType,
    DetectorModel, DetectorType, Manufacturer, Supplier,
    SensorType, Sensor, SensorSlot, SensorGas, SensorStatus,
    DetectorModelConfiguration,
    Detector, DetectorStatus,
    Maintenance, MaintenanceType, MaintenanceStatus,
    MaintenanceTask, MaintenanceTaskType,
    Cylinder, CylinderType, CylinderStatus,
    DetectorFault,
)


# Reverse mapping dictionaries (code -> description)
LOCATION_TYPE_MAP = {v: k for k, v in LocationType.choices}
MANUFACTURER_MAP = {v: k for k, v in Manufacturer.choices}
SUPPLIER_MAP = {v: k for k, v in Supplier.choices}
DETECTOR_TYPE_MAP = {v: k for k, v in DetectorType.choices}
DETECTOR_STATUS_MAP = {v: k for k, v in DetectorStatus.choices}
MAINTENANCE_TYPE_MAP = {v: k for k, v in MaintenanceType.choices}
SENSOR_GAS_MAP = {v: k for k, v in SensorGas.choices}
CYLINDER_GAS_MAP = {v: k for k, v in CylinderStatus.choices}
SENSOR_STATUS_MAP = {v: k for k, v in SensorStatus.choices}


def get_thin_border():
    """Return a thin border style for cells."""
    return Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin'),
    )


def style_sheet(ws, headers):
    """Apply basic styling to a worksheet."""
    # Style header row
    header_font = Font(bold=True)
    header_alignment = Alignment(horizontal='center', vertical='center')
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = get_thin_border()
    
    # Auto-adjust column widths
    for i, header in enumerate(headers, 1):
        if header:
            ws.column_dimensions[get_column_letter(i)].width = min(len(str(header)) + 2, 25)
        else:
            ws.column_dimensions[get_column_letter(i)].width = 15


def create_lookups_sheet(wb):
    """Create the Lookups sheet with all choice mappings."""
    print("Creating Lookups sheet...")
    ws = wb.active
    ws.title = 'Lookups'
    
    # Define lookup tables
    lookups = [
        ('LocationType', LocationType.choices),
        ('Manufacturer', Manufacturer.choices),
        ('DetectorType', DetectorType.choices),
        ('Supplier', Supplier.choices),
        ('DetectorStatus', DetectorStatus.choices),
        ('MaintenanceType', MaintenanceType.choices),
        ('MaintenanceTaskType', MaintenanceTaskType.choices),
        ('MaintenanceStatus', MaintenanceStatus.choices),
        ('DetectorFaultStatus', [('Open', 'OP'), ('Closed', 'CL')]),
        ('DetectorFaultType', [('Failed Bump', 'BF'), ('Sensor Fail', 'SF'), ('Displays Error', 'DE'), 
                               ('Will not turn on', 'WS'), ('Damaged Display', 'DD'), ('Damaged Casing', 'DC'),
                               ('Missing Attachment', 'MA')]),
        ('CylinderGas', [('CO', 'CO'), ('H2S', 'HS'), ('CH4', 'CH'), ('O2', 'O2'), ('Iso', 'IB'),
                         ('HCN', 'HC'), ('N2', 'N2'), ('Cl2', 'CL'), ('PH3', 'PH'), ('SO2', 'SO'),
                         ('NO2', 'NO'), ('CO2', 'C2'), ('NH3', 'NH'), ('ETO', 'ET')]),
        ('SensorGas', SensorGas.choices),
        ('CylinderVolume', [('34 L', 'L034'), ('65 L', 'L065'), ('103 L', 'L103'), 
                            ('112 L', 'L112'), ('552 L', 'L552')]),
        ('CylinderUnit', [('ppm', 'PM'), ('%v/v', 'PV'), ('%LEL', 'PL'), ('mg/L', 'ML')]),
        ('CylinderStatus', [('On Order', 'OO'), ('In Stock', 'IS'), ('Operational', 'OP'), ('Empty', 'MT')]),
        ('SensorStatus', SensorStatus.choices),
    ]
    
    # Build header row
    headers = []
    for name, _ in lookups:
        headers.extend([name, 'Code'])
    
    # Write headers
    for col_num, header in enumerate(headers, 1):
        ws.cell(row=1, column=col_num, value=header)
    
    # Find max rows needed
    max_rows = max(len(choices) for _, choices in lookups)
    
    # Write data
    for row_idx in range(max_rows):
        col_num = 1
        for _, choices in lookups:
            if row_idx < len(choices):
                label, code = choices[row_idx]
                ws.cell(row=row_idx + 2, column=col_num, value=label)
                ws.cell(row=row_idx + 2, column=col_num + 1, value=code)
            col_num += 2
    
    style_sheet(ws, headers)
    print(f"  Created Lookups sheet with {len(lookups)} lookup tables")


def create_location_sheet(wb):
    """Create the Location sheet."""
    print("Creating Location sheet...")
    ws = wb.create_sheet('Location')
    
    headers = ['label', 'station', 'address', 'location_type_desc', 'location_type_code', 'priority']
    
    locations = Location.objects.all().order_by('location_type', 'label')
    
    # Write headers
    for col_num, header in enumerate(headers, 1):
        ws.cell(row=1, column=col_num, value=header)
    
    # Write data
    for row_idx, location in enumerate(locations, 2):
        ws.cell(row=row_idx, column=1, value=location.label)
        ws.cell(row=row_idx, column=2, value=location.station)
        ws.cell(row=row_idx, column=3, value=location.address)
        location_type_desc = LOCATION_TYPE_MAP.get(location.location_type, location.location_type)
        ws.cell(row=row_idx, column=4, value=location_type_desc)
        ws.cell(row=row_idx, column=5, value=location.location_type)
        ws.cell(row=row_idx, column=6, value=location.priority)
    
    style_sheet(ws, headers)
    print(f"  Created Location sheet with {locations.count()} locations")


def create_detector_models_sheet(wb):
    """Create the DetectorModels sheet."""
    print("Creating DetectorModels sheet...")
    ws = wb.create_sheet('DetectorModels')
    
    headers = ['model_name', 'detector_type_desc', 'detector_type_code', 
               'manufacturer_desc', 'manufacturer_code', 'supplier_desc', 'supplier_code']
    
    detector_models = DetectorModel.objects.all().order_by('detector_type', 'label')
    
    # Write headers
    for col_num, header in enumerate(headers, 1):
        ws.cell(row=1, column=col_num, value=header)
    
    # Write data
    for row_idx, model in enumerate(detector_models, 2):
        ws.cell(row=row_idx, column=1, value=model.label)
        detector_type_desc = DETECTOR_TYPE_MAP.get(model.detector_type, model.detector_type)
        ws.cell(row=row_idx, column=2, value=detector_type_desc)
        ws.cell(row=row_idx, column=3, value=model.detector_type)
        manufacturer_desc = MANUFACTURER_MAP.get(model.manufacturer, model.manufacturer) if model.manufacturer else ''
        ws.cell(row=row_idx, column=4, value=manufacturer_desc)
        ws.cell(row=row_idx, column=5, value=model.manufacturer or '')
        supplier_desc = SUPPLIER_MAP.get(model.supplier, model.supplier) if model.supplier else ''
        ws.cell(row=row_idx, column=6, value=supplier_desc)
        ws.cell(row=row_idx, column=7, value=model.supplier or '')
    
    style_sheet(ws, headers)
    print(f"  Created DetectorModels sheet with {detector_models.count()} models")


def create_sensor_types_sheet(wb):
    """Create the SensorTypes sheet."""
    print("Creating SensorTypes sheet...")
    ws = wb.create_sheet('SensorTypes')
    
    headers = ['sensorgas', 'part_number', 'manufacturer', 'compatible_detectormodels', 
               'warranty_months', 'expiry_months']
    
    sensor_types = SensorType.objects.all().order_by('part_number')
    
    # Write headers
    for col_num, header in enumerate(headers, 1):
        ws.cell(row=1, column=col_num, value=header)
    
    # Write data
    for row_idx, sensor_type in enumerate(sensor_types, 2):
        sensor_gas = SENSOR_GAS_MAP.get(sensor_type.sensorgas, sensor_type.sensorgas)
        ws.cell(row=row_idx, column=1, value=sensor_gas)
        ws.cell(row=row_idx, column=2, value=sensor_type.part_number)
        manufacturer = MANUFACTURER_MAP.get(sensor_type.manufacturer, sensor_type.manufacturer) if sensor_type.manufacturer else ''
        ws.cell(row=row_idx, column=3, value=manufacturer)
        ws.cell(row=row_idx, column=4, value=sensor_type.compatible_detectormodels)
        ws.cell(row=row_idx, column=5, value=sensor_type.warranty_months)
        ws.cell(row=row_idx, column=6, value=sensor_type.expiry_months)
    
    style_sheet(ws, headers)
    print(f"  Created SensorTypes sheet with {sensor_types.count()} sensor types")


def create_detector_configurations_sheet(wb):
    """Create the DetectorConfigurations sheet."""
    print("Creating DetectorConfigurations sheet...")
    ws = wb.create_sheet('DetectorConfigurations')
    
    headers = ['label', 'detector_type', 'sensor_gases']
    
    configurations = DetectorModelConfiguration.objects.all().select_related('detector_model').order_by('detector_model__label', 'label')
    
    # Write headers
    for col_num, header in enumerate(headers, 1):
        ws.cell(row=1, column=col_num, value=header)
    
    # Write data
    for row_idx, config in enumerate(configurations, 2):
        ws.cell(row=row_idx, column=1, value=config.label)
        ws.cell(row=row_idx, column=2, value=config.detector_model.label)
        
        # Convert sensor gas codes back to descriptions
        sensor_gases_desc = ''
        if config.sensor_gases:
            gas_codes = config.sensor_gases.split(',')
            gas_descriptions = [SENSOR_GAS_MAP.get(code, code) for code in gas_codes]
            sensor_gases_desc = ','.join(gas_descriptions)
        ws.cell(row=row_idx, column=3, value=sensor_gases_desc)
    
    style_sheet(ws, headers)
    print(f"  Created DetectorConfigurations sheet with {configurations.count()} configurations")


def create_detectors_sheet(wb):
    """Create the Detectors sheet."""
    print("Creating Detectors sheet...")
    ws = wb.create_sheet('Detectors')
    
    headers = ['label', 'serial', 'detector_model__label', 'location__label', 
               'configuration', 'status']
    
    detectors = Detector.objects.all().select_related(
        'detector_model', 'location', 'configuration'
    ).order_by('label')
    
    # Write headers
    for col_num, header in enumerate(headers, 1):
        ws.cell(row=1, column=col_num, value=header)
    
    # Write data
    for row_idx, detector in enumerate(detectors, 2):
        ws.cell(row=row_idx, column=1, value=detector.label)
        ws.cell(row=row_idx, column=2, value=detector.serial)
        ws.cell(row=row_idx, column=3, value=detector.detector_model.label)
        ws.cell(row=row_idx, column=4, value=detector.location.label)
        ws.cell(row=row_idx, column=5, value=detector.configuration.label if detector.configuration else '')
        status_desc = DETECTOR_STATUS_MAP.get(detector.status, detector.status)
        ws.cell(row=row_idx, column=6, value=status_desc)
    
    style_sheet(ws, headers)
    print(f"  Created Detectors sheet with {detectors.count()} detectors")


def create_maintenance_sheet(wb):
    """Create the Maintenance sheet."""
    print("Creating Maintenance sheet...")
    ws = wb.create_sheet('Maintenance')
    
    headers = ['label', 'maintenance_type', 'due_date']
    
    maintenances = Maintenance.objects.all().select_related('detector').order_by('detector__label', 'date_due')
    
    # Write headers
    for col_num, header in enumerate(headers, 1):
        ws.cell(row=1, column=col_num, value=header)
    
    # Write data
    for row_idx, maintenance in enumerate(maintenances, 2):
        ws.cell(row=row_idx, column=1, value=maintenance.detector.label)
        maintenance_type_desc = MAINTENANCE_TYPE_MAP.get(maintenance.maintenance_type, maintenance.maintenance_type)
        ws.cell(row=row_idx, column=2, value=maintenance_type_desc)
        ws.cell(row=row_idx, column=3, value=maintenance.date_due)
    
    style_sheet(ws, headers)
    print(f"  Created Maintenance sheet with {maintenances.count()} maintenance records")


def create_sensors_sheet(wb):
    """Create the Sensors sheet."""
    print("Creating Sensors sheet...")
    ws = wb.create_sheet('Sensors')
    
    headers = ['detector', 'serial', 'part_number', 'manufacture_date', 'warranty_date', 'expiry_date']
    
    sensors = Sensor.objects.all().select_related('sensor_type', 'detector').order_by('sensor_type__part_number', 'serial')
    
    # Write headers
    for col_num, header in enumerate(headers, 1):
        ws.cell(row=1, column=col_num, value=header)
    
    # Write data
    for row_idx, sensor in enumerate(sensors, 2):
        ws.cell(row=row_idx, column=1, value=sensor.detector.label if sensor.detector else '')
        ws.cell(row=row_idx, column=2, value=sensor.serial or '')
        ws.cell(row=row_idx, column=3, value=sensor.sensor_type.part_number)
        ws.cell(row=row_idx, column=4, value=sensor.receive_date)  # receive_date is manufacture_date
        ws.cell(row=row_idx, column=5, value=sensor.warranty_date)
        ws.cell(row=row_idx, column=6, value=sensor.expiry_date)
    
    style_sheet(ws, headers)
    print(f"  Created Sensors sheet with {sensors.count()} sensors")


def main():
    """Main function to export the database to Excel."""
    print("=" * 60)
    print("Exporting Inventory Database to Excel")
    print("=" * 60)
    
    # Create workbook
    wb = Workbook()
    
    # Create sheets in order
    print("\nCreating worksheets...")
    create_lookups_sheet(wb)
    create_location_sheet(wb)
    create_detector_models_sheet(wb)
    create_sensor_types_sheet(wb)
    create_detector_configurations_sheet(wb)
    create_detectors_sheet(wb)
    create_maintenance_sheet(wb)
    create_sensors_sheet(wb)
    
    # Remove default sheet if it exists
    if 'Sheet' in wb.sheetnames:
        del wb['Sheet']
    
    # Save workbook
    output_path = os.path.join(BASE_DIR, 'inventory', 'fixtures', 'Equipment List Export.xlsx')
    print(f"\nSaving to: {output_path}")
    wb.save(output_path)
    
    print("\n" + "=" * 60)
    print("Export complete!")
    print("=" * 60)
    print(f"\nSummary:")
    print(f"  Locations: {Location.objects.count()}")
    print(f"  Detector Models: {DetectorModel.objects.count()}")
    print(f"  Sensor Types: {SensorType.objects.count()}")
    print(f"  Detector Configurations: {DetectorModelConfiguration.objects.count()}")
    print(f"  Detectors: {Detector.objects.count()}")
    print(f"  Maintenance Records: {Maintenance.objects.count()}")
    print(f"  Sensors: {Sensor.objects.count()}")
    print(f"\nOutput file: {output_path}")


if __name__ == '__main__':
    main()
